from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

wb = Workbook()
ws_entrada = wb.active
ws_entrada.title = 'Sheet1'

cabecalhos_entrada = [
    'material',
    'ncm da nota',
    'cadastro sugerido',
    'codigo existente',
    'ncm sugerido',
    'tipo sugerido',
    'STATUS',
]

dados_entrada = [
    ['Bomb. pneum s/bc pt', '', '', '', '', '', ''],
    ['tubo galv 1/2', '', '', '', '', '', ''],
    ['paraf sext 1/4 x 20', '', '', '', '', '', ''],
]

header_fill = PatternFill('solid', fgColor='4472C4')
header_font = Font(bold=True, color='FFFFFF')
thin = Side(style='thin')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for col, cabecalho in enumerate(cabecalhos_entrada, start=1):
    cell = ws_entrada.cell(row=1, column=col, value=cabecalho)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')
    cell.border = border
    ws_entrada.column_dimensions[cell.column_letter].width = 22

for row_idx, linha in enumerate(dados_entrada, start=2):
    for col_idx, valor in enumerate(linha, start=1):
        cell = ws_entrada.cell(row=row_idx, column=col_idx, value=valor)
        cell.border = border
        if col_idx == 1:
            cell.fill = PatternFill('solid', fgColor='FFF2CC')

ws_banco = wb.create_sheet('banco')
cabecalhos_banco = ['codigo', 'material', 'ncm', 'tipo']
dados_banco = [
    ['001234', 'BOMBA PNEUMATICA SEM BOCA PRETA', '84137010', 'UN'],
    ['005678', 'TUBO GALVANIZADO 1 POLEGADA', '73063000', 'MT'],
]

for col, cabecalho in enumerate(cabecalhos_banco, start=1):
    cell = ws_banco.cell(row=1, column=col, value=cabecalho)
    cell.fill = PatternFill('solid', fgColor='548235')
    cell.font = Font(bold=True, color='FFFFFF')
    cell.border = border

for row_idx, linha in enumerate(dados_banco, start=2):
    for col_idx, valor in enumerate(linha, start=1):
        ws_banco.cell(row=row_idx, column=col_idx, value=valor).border = border

wb.save('planilha.xlsx')
print('planilha.xlsx criada com formatação de exemplo.')
